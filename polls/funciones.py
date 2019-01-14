#!/usr/bin/env python
# -*- coding: cp1252 -*-
# -*- coding: 850 -*-
# -*- coding: utf-8 -*-
import configparser
import os
import random
import sqlite3
import ssl
import urllib
import re
import urllib2
import getpass
import requests
from bs4 import BeautifulSoup
from mechanize import Browser
import time
import datetime
from PIL import ImageTk, Image
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
#Esta funcion te li
def limpiarNombre(nombre):
	acum=""
	ordenamiento=0
	for elemento in nombre:
		ordenamiento=ord(elemento)
		if ordenamiento >=65 and ordenamiento <=122 or ordenamiento==32:
			acum+=elemento
		else:
			if ordenamiento == 209:
				acum+="Ñ"
			else:
				acum+="-"
	return acum
def chequearSPINespecial(documento):
    print "USANDO SPIN ESPECIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAL----------------------------"
    global bateria
    global porta
    global nombreglobal2
    url = configPOS['Superintendencia']['Url']
    nombreglobal2=""
    cadenita2 = ""
    cadenita = "" 
    bateria = ""
    franco = False
    encontro = False 
    encontro2 = False
    browser.open(url,timeout=2)
    form = browser.select_form(nr=0)
    browser["nro_doc"]=documento
    response =browser.submit()
    contadorsito2=0
    contadorsitocade2=0
    OSSI2=""
    leer = response.read()
    for elemento in leer:
        if encontro2 == True and elemento == "<":
            if contadorsitocade2 == 6:
                nombreglobal2 = cadenita2
            encontro2=False
            cadenita2=""
            contadorsitocade2+=1
        if encontro == True and elemento == "<":
            contadorsito2+=1
            if cadenita.strip(" ") == "No se reportan datos para el NUMERO DE DOCUMENTO "+documento:
                OSSI2="No se encontraron datos"
                return OSSI2
            if contadorsito2 == 3:
                OSSI2=cadenita
            if contadorsito2 == 4:
                bateria=cadenita
            encontro=False
            cadenita=""
        if elemento == "<":
            cadenita2=""
            cadenita=""
        cadenita2+=elemento
        cadenita+=elemento
        if cadenita2== '<td>':
            encontro2=True
            cadenita2=""
        if cadenita == '<b>':
            encontro=True
        if elemento==">":
            cadenita=""
    return OSSI2
#ControlSPIN controla que funcion usar dependiendo si el bot ya esta logeado en la pagina o no
def controlSPIN(documento):
    try:
        ssl._create_default_https_context = ssl._create_unverified_context
        url = "https://seguro.sssalud.gob.ar/indexss.php?opc=bus650&user=HPGD&cat=consultas"
        header = {'User-Agent': 'Mozilla/5.0 (Windows NT 5.1; rv:14.0) Gecko/20100101 Firefox/14.0.1', 'Referer': 'http://whateveritis.com'}
        request = urllib2.Request(url, None, header)
        response = browser.open(request,timeout=2)
        form = browser.select_form(nr=0)
        try:
            browser["_user_name_"] =usuarioL
            browser["_pass_word_"] =contraL
            response =browser.submit()
            OSSI2=chequearSPIN(documento,response)
            return OSSI2
        except:
        	#Si entro aca significa que el bot ya estaba logeado
            OSSI2=chequearSPINespecial(documento)
            return OSSI2
    except Exception as e:
        print "----------------------------------------------------------------------------------------"
        print e
        print "----------------------------------------------------------------------------------------"
        return "ERROR"      
def chequearSPIN(documento,response):
    global bateria
    global porta
    global nombreglobal2
    nombreglobal2=""
    cadenita2 = ""
    cadenita = "" 
    bateria = ""
    franco = False
    encontro = False 
    encontro2 = False
    form = browser.select_form(nr=0)
    browser["nro_doc"]=documento
    response =browser.submit()
    contadorsito2=0
    contadorsitocade2=0
    OSSI2=""
    leer = response.read()
    #TODO EL FOR ESTE, PARSEA EL HTML PARA CONSEGUIR LA INFORMACION QUE NECESESITAMOS
    for elemento in leer:
        if encontro2 == True and elemento == "<":
            if contadorsitocade2 == 6:
                nombreglobal2 = cadenita2
            encontro2=False
            cadenita2=""
            contadorsitocade2+=1
        if encontro == True and elemento == "<":
            contadorsito2+=1
            if cadenita.strip(" ") == "No se reportan datos para el NUMERO DE DOCUMENTO "+documento:
                OSSI2="No se encontraron datos"
                return OSSI2
            if contadorsito2 ==3:
                OSSI2=cadenita
            if contadorsito2 == 4:
                bateria=cadenita
            encontro=False
            cadenita=""
        if elemento == "<":
            cadenita2=""
            cadenita=""
        cadenita2+=elemento
        cadenita+=elemento
        if cadenita2== '<td>':
            encontro2=True
            cadenita2=""
        if cadenita == '<b>':
            encontro=True
        if elemento==">":
            cadenita=""
    return OSSI2
def chequearIOMA2018(documento):
	tiene=False
	#ESTE FOR ESTA PARA ITERAR SOBRE LOS SEXOS (1,2) ESTO SE HACE POR QUE PARA CONSULTAR A IOMA HAY QUE ELEGIR EL SEXO
	for item in range(1,3):
		if chequearIOMA2018bis(documento,item) == True:
			return True
	return False
def chequearIOMA2018bis(documento,sexo):
	#HACE UNA REQUEST A LA PAGINA DE IOMA CON LOS HEADERS EN LA VARIABLE 'DATA' DE ABAJO
	data = urllib.urlencode({'T3':documento,'sexo':sexo,'B13':'Buscar'})
	url=configPOS['Ioma']['Url']
	request = urllib2.Request(url,data)
	respuesta = urllib2.urlopen(request).read()
	counterIOMA=0
	itworks=False
	suma=0
	cadenita = ""
	ultima = ""
	franco = False
	encontro = False
	#PARSEA LA RESPUESTA QUE SERIA UN HTML
	for elemento in respuesta:
		if encontro == True and elemento == "<":
			counterIOMA+=1
			if counterIOMA == 2:
				nombreglobal3 = cadenita
			ultima=cadenita
			encontro=False
			cadenita=""
		if elemento == "<":
			cadenita=""
		cadenita+=elemento
		if cadenita == '<span class="texto-azul">' or cadenita =='<span class="texto-azul-bold">' :
			encontro=True
			itworks = True
		if elemento==">":
			cadenita=""
	if len(ultima)>2:
		return True
	else:
		return False
def limpiandingSTR(cadena):
    nombresito=""
    c=""
    nombre=""
    listreturn=[]
    conta=1
    conta2=0
    eselnombre=False
    encontro = False
    termino= False
    nueva=0
    for elemento in cadena:
        if eselnombre == True:
            if elemento == '"':
                nombresito=nombre
                eselnombre=False
        nombre+=elemento
        if elemento == ",":
            nombre=""
        if elemento != '"' and elemento != "}" and encontro == True: 
            c+=elemento
        if elemento == ":":
            c=""
            encontro =True
        if elemento == "}":
            if c not in listreturn and len(c)>3:
                listreturn.append(c)
            conta2=0
        if nombre == '"NombreYApellido":"':
            eselnombre=True
            nombre=""
    return listreturn,nombresito
def limpiandingSTR2(cadena):
    global estaactivosumar
    clausula = False
    clausula2 = False
    clausula3= False
    apellido=""
    nombreactivo=""
    eselapellido=False
    nombresito=""
    nombresito2=""
    c=""
    nombre=""
    listreturn=[]
    conta=1
    conta2=0
    eselnombre=False
    encontro = False
    termino= False
    nueva=0
    for elemento in cadena:
        if eselnombre == True:
            if elemento == '"' and clausula == False:
                nombresito=nombre
                eselnombre=False
        if eselapellido == True:
            if elemento == '"' and clausula2 == False:
                nombresito2=apellido
                eselapellido=False
        apellido+=elemento
        nombre+=elemento
        nombreactivo+=elemento
        if elemento == "," and encontro == True:
            nombre=""
        if elemento != '"' and elemento != "}" and encontro == True: 
            c+=elemento
        if elemento == ":":
            c=""
            encontro =True
        if elemento == "}":
            listreturn.append(c)
            conta2=0
        if nombre == '"afiNombre":"':
            eselnombre=True
            nombre=""
        if nombre == '"afiApellido":"':
            eselapellido=True
            apellido=""
        # if nombre == '"Activo":"N"' and clausula3 == False:
        #   clausula3= True
        #   estaactivosumar=False
        if nombre == '"Activo":"S"' and clausula3 == False:
            estaactivosumar = True
            return nombresito+" "+nombresito2
    estaactivosumar = False
    return ""
#CONSEGUIRINPUT LEE TODA LA REQUEST Y TRAE LA VARIABLE PUESTA EN EL INPUT DEL TEMPLATE 'ConsultaUnica.html' TOTALMENTE MAL, DEBE HABER ALGO ULTRA SIMPLE PERO ME DIO PAJA BUSCARLO (algo como request.input)
def conseguirInput(cadena):
    inputLimpio=""
    corchetes=0
    encontroInput=False
    pasoLos2Car=False
    conta=0
    for elemento in cadena:
        if encontroInput==True and conta == 2 and elemento=="'":
            return inputLimpio
        if elemento == "[" and encontroInput==False:
            corchetes+=1
        if encontroInput==True and conta<2:
            conta+=1
        if corchetes==2:
            encontroInput=True
        if conta==2:
            if elemento!="'":
                inputLimpio+=elemento
def siLaPalabraEstaEnLaLista(lista,palabra):
    for elemento in lista:
        if palabra in elemento:
            return True
    return False
def tienePAMI(laStr, elemento1, elemento2):
    if laStr in elemento1 or siLaPalabraEstaEnLaLista(elemento2,laStr):
        return True
    return False
def chequearPAMI(documento):
    browser.open('http://institucional.pami.org.ar/result.php?c=6-2')
    form=browser.select_form(nr=1)
    browser.set_handle_robots( False )
    browser["nroDocumento"]=documento
    response =browser.submit()
    respuesta=response.read()
    return filtrarVomitoV2(respuesta)
def filtrarVomito(vomito):
    url=""
    acum=""
    acumLink=""
    encontroLink=False
    loproximo=False
    lodeahora=False
    for elemento in vomito:
        if lodeahora==True: 
            if elemento !='"':
                acumLink+=elemento
            else:
                if "beneficio" in acumLink and "parent" in acumLink and encontroLink == False:
                    encontroLink=True
                    url=acumLink
                lodeahora=False
                acumLink=""
        if elemento == "<":
            if loproximo==True:
                loproximo=False
        acum+=elemento
        if elemento == ">":
            if acum == '<p class="whitetxt">':
                loproximo=True
            acum=""
        if acum=='<a href="':
            lodeahora=True
    url="http://institucional.pami.org.ar/"+url
    return chequearArbolPami(url)
def filtrarVomitoV2(vomito):
    global contaRAROS
    listaLinks=[]
    listaNom=[]
    listaAlta=[]
    listaBaja=[]
    listaDoc=[]
    listadoPami=[]
    url=""
    acum=""
    acumLink=""
    encontroLink=False
    loproximo=False
    lodeahora=False
    contaRAROS=0
    lista=["td"]
    cunta=0
    soup = BeautifulSoup(vomito,'html.parser')
    for elemento in soup.findAll('p', attrs={'class': 'whitetxt'}):
        cunta+=1
        # print elemento.text
        if cunta==1:
            listaNom.append(elemento.text)
        if cunta==4:
            listaDoc.append(elemento.text)
        if cunta==5:
            listaAlta.append(elemento.text)
        if cunta%6==0:
            listaBaja.append(elemento.text)
            # raw_input()
            cunta=0
    contaLetras=""
    for link in soup.findAll('a'):
        if link.string=="add":
            for elemento in str(link):
                contaLetras+=elemento 
                if elemento == '"':
                    if "href" in contaLetras:
                        contaLetras=""
                    if "beneficio" in contaLetras and "parent" in contaLetras:
                        listaLinks.append(contaLetras.rstrip('"').replace("amp;", ''))
                        contaLetras=""
    cuntaLE=0
    listado1=[]
    listado2=[]
    listado3=[]
    for elemento in listaLinks:
        if len(listaBaja[cuntaLE]) < 2:
            print listaNom[cuntaLE]
            print listaDoc[cuntaLE]
            print listaAlta[cuntaLE]
            print listaBaja[cuntaLE]
            print listaLinks[cuntaLE]
            url="http://institucional.pami.org.ar/"+listaLinks[cuntaLE]
            listado1,listado2,listado3= chequearArbolPami(url)
            listadoPami.append([listaDoc[cuntaLE],listaNom[cuntaLE]])
            listadoPami.append(zip(listado1,listado2,listado3))
        cuntaLE+=1
    cuntaLE=0
    return listadoPami
#ESTA FUNCION TE DEVUELVE LAS 3 LISTAS DE PAMI
def chequearArbolPami(url):
    r  = requests.get(url)
    listaDmodulo=[]
    listaRed=[]
    listaPrestador=[]
    contador=0
    contaTabla=1
    data = r.text
    soup = BeautifulSoup(data,"html5lib")
    for link in soup.find_all('p'):
        # print(link.get('class'))
        if "<p>PRESTADOR:</p>" in str(link): 
            contador+=1
        if contador==2 and "<p>PRESTADOR:</p>" not in str(link):
            linkLimpio=str(link).strip("<p>").strip("</p>")
            if "APELLIDOS Y NOMBRES" in str(link):
                break
            if contaTabla==4:
                contaTabla=1
            if contaTabla==1:
                # print "D.MODULO: "+linkLimpio
                listaDmodulo.append(str(linkLimpio))
            if contaTabla==2:
                # print "RED: "+linkLimpio
                listaRed.append(str(linkLimpio))
            if contaTabla==3:
                # print "PRESTADOR: "+linkLimpio
                listaPrestador.append(str(linkLimpio))
            contaTabla+=1
    return listaDmodulo,listaRed,listaPrestador
#BUSCA ENTRE TODOS LOS NOMBRES QUE TIENE DE LOS BOTS Y TE DEVUELVE UNO VALIDO
def nombreLigadoAlDocumento():
    if len(nombreglobal)>4:
        return nombreglobal
    if len(nombreglobal3)>4:
        return nombreglobal3
    if len(nombreglobal2)>4:
        return nombreglobal2
    return False
def tieneOS(IOMA,PUCO,SPIN):
	tiene=False
	if IOMA != False and IOMA != "ERROR":
		return True
	if SPIN != "No se encontraron datos" and SPIN != "ERROR":
		return True 
	if PUCO != ["No se reportan datos"] and len(PUCO[0]) > 1:
		return True
	return tiene
def limpiarNombres():
	global nombreglobalsumar
	global nombreglobal
	global nombreglobal2
	global nombreglobal3
	nombreglobalsumar=""
	nombreglobal=""
	nombreglobal2=""
	nombreglobal3=""
def escribirEnLaBD(documento,nombre,puco,ioma,spin):
   puco2=""
   nombre=nombre.strip("- ")
   spin=spin.strip(" ")
   if type(puco) == list:
       for elemento in puco:
           puco2+=elemento+", "
       puco=puco2
   conn = sqlite3.connect(os.path.join( os.path.dirname( __file__),'..')+"/static/pos.sqlite3") 
   cursor = conn.cursor()
   conn.text_factory = str
   cursor.execute('''INSERT INTO historial(documento,nombre,puco,ioma,spin,fecha) VALUES(:documento, :nombre, :puco, :ioma, :spin, :fecha)''',                  {'documento':documento, 'nombre':nombre, 'puco':puco, 'ioma':ioma, 'spin':spin, 'fecha':datetime.datetime.now() })
   conn.commit()
   cursor.execute('SELECT * FROM historial')
   data = cursor.fetchall()
def sacarEstosElementos(cadena,lista):
    for elemento in lista:
        cadena=cadena.strip(elemento)
    return cadena
def botPUCO2018lista(listaDoc,listaReal):
	global intentoListaDoc
	global nombreglobal
	listaDocumentos=[]
	listaObrasSociales=[]
	listaNombres=[]
	par=0
	listaBan=[",",":"]
	nombreglobal = ""
	data = urllib.urlencode({'documento':listaDoc,'tabla':'PUCO_2018-12'})
	try:
		url = configPOS['Puco']['UrlPrincipal']
		request = urllib2.Request(url,data)
		respuesta = urllib2.urlopen(request).read()
		regex = r"\{(.*?)\"}"
		elDocumento=""
		esElDoc=False
		ultimo=""
		elNombre=""
		matches = re.finditer(regex, respuesta, re.MULTILINE | re.DOTALL)
		for matchNum, match in enumerate(matches):
			par=0
			for groupNum in range(0, len(match.groups())):       
				elmo=match.group(1)
				esElDoc=False
				for elemento in elmo.split('"'):
					if elemento not in listaBan:
						# print elemento
						if par%2==0:
							if par==4:
								elDocumento=elemento
							if par==6:
								elNombre=elemento
                            # print "--------------------------------------------------------------"
						ultimo=elemento

						par+=1
                # print "******************************************"
				listaDocumentos.append(elDocumento)
				listaNombres.append(elNombre)
				listaObrasSociales.append(ultimo)
		lezipeado=zip(listaDocumentos,listaNombres,listaObrasSociales)
		return lezipeado
	except Exception as e:
		print e 
		listaEr=["PAGINA NO DISPONIBLE"]
		return listaEr
	return listaDocumentos,listaNombres,listaObrasSociales
def nombreSumar():
	global nombreglobalsumar
	return (nombreglobalsumar)
def botPUCOSUMAR(documento):
    try:
        global nombreglobalsumar
        data = urllib.urlencode({'documento':documento})
        url = configPOS['Puco']['UrlSumar']
        request = urllib2.Request(url,data)
        respuesta = urllib2.urlopen(request).read()
        nombreglobalsumar=limpiandingSTR2(respuesta)
        if respuesta == "null" or estaactivosumar == False:
            return False
        else:
            return True
    except:
        return "ERROR"
#LE DAS UNA LISTA Y TE LO TRANSFORMA EN UN CSV 
def convertirEnUna(masiva):
    listaDOC=[]
    listaOS=[]
    listaNOM=[]
    anterior=0
    acum=0
    for doc,os,nom in masiva:
        if anterior==doc:
            listaOS[acum-1]=listaOS[acum-1]+" ; "+os
        else:
            listaDOC.append(doc)
            listaOS.append(os)
            listaNOM.append(nom)
            acum+=1
        anterior=doc
    return zip(listaDOC,listaOS,listaNOM)
#LISTADOCUS2018 LEE TODOS LOS DOCUMENTOS DEL EXCEL Y TE LOS DEVUELVE EN UNA LISTA
def listaDocus2018bis(hojaDocus):
    columnaDocus="A"
    n=1
    listaDocus=[]
    contaNones=0
    for fila in hojaDocus:
        if contaNones>9:
                break
        for columna in fila:
            if contaNones>9:
                break
            if columna.coordinate==columnaDocus+str(n):
                if columna.value==None or len(str(columna.value))<7 and columna.value>1000000:
                    contaNones+=1
                else:
                    contaNones=0
                    listaDocus.append(str(columna.value))
        n+=1
    return listaDocus
def botPUCO2018listaBis(lista):
	listaLimpia=[]
	acum=0
	listaDoc=[]
	listaNom=[]
	listaOSES=[]
	listaDoc1=[]
	listaNom1=[]
	listaOSES1=[]
	registro1=""
	registro2=""
	nidea=""
	listadocus=""
	for elemento in lista:
		listadocus+=filtroSoloNumeros(elemento)+" "
		if acum%2500==0 and acum>1 or acum+1==len(lista):
			niidea=botPUCO2018lista(listadocus.strip(" "),lista)
			for a,b,c in niidea:
				listaDoc.append(a)
				listaNom.append(b)
				listaOSES.append(c)
			listadocus=""
		acum+=1
	acum=0
	acum2=0
	return filtroDeListongui(listaDoc,listaOSES,listaNom)
def filtroSoloNumeros(numero):
    acum=""
    if numero != None:
        for elemento in numero:
            if elemento.isdigit() == True:
                acum+=elemento
        return acum
#ESCRIBE EN UN EXCEL TODOS LOS DATOS DE PUCO 
def listaDocus2018(archivo):
    hojaDocumentos=archivo.active
    listaTerminada=botPUCO2018listaBis(listaDocus2018bis(hojaDocumentos))
    print type(listaTerminada)
    vacio=Workbook()
    hojaDocumentos=vacio.active
    n=1
    hojaDocumentos.column_dimensions["C"].width=75
    hojaDocumentos.column_dimensions["B"].width=35
    for a,b,c in listaTerminada:
        if c == "NO":
            hojaDocumentos["A"+str(n)].value=a
            hojaDocumentos["B"+str(n)].value="NO HAY DATOS"
            hojaDocumentos["C"+str(n)].value="NO HAY DATOS"
        else:
            hojaDocumentos["A"+str(n)].value=a
            hojaDocumentos["B"+str(n)].value=c
            hojaDocumentos["C"+str(n)].value=b
        n+=1
    return vacio
def filtroDeListongui(listaDocus,listaOses,listaNom):
    conta=0
    anterior=""
    todo=zip(listaDocus,listaOses,listaNom) 
    while conta<=len(listaDocus)-1:
        if anterior==listaDocus[conta]:
            listaOses[conta-1]=listaOses[conta-1]+" ; "+listaOses[conta]
            listaOses.pop(conta)
            listaDocus.pop(conta)
            listaNom.pop(conta)
            conta=conta-1
        anterior=listaDocus[conta]
        conta+=1
    return zip(listaDocus,listaOses,listaNom)
def chequearPAMIconTIPO(documento):
    listaOpciones=["DNI","LC","LE","PAS","CI"]
    vomitoLimpio=[]
    listaZips=[]
    listaTipo=[]
    for elemento in listaOpciones:
        print elemento
        browser.open('http://institucional.pami.org.ar/result.php?c=6-2')
        form=browser.select_form(nr=1)
        browser.set_handle_robots( False )
        browser["nroDocumento"]=documento
        browser["tipoDocumento"]=[elemento]
        response =browser.submit()
        respuesta=response.read()
        lista1a,lista2b,lista3b=filtrarVomito(respuesta)
        vomitoLimpio=zip(lista1a,lista2b,lista3b)
        if len(vomitoLimpio)>1 and len(vomitoLimpio[0][0])>0:
            listaZips.append(elemento)
            listaZips.append(vomitoLimpio)

    acum=0
    return listaZips
def masInformacionBots(documento):
    listaZips=[]
    listaTipos=[]
    sumar=""
    listaZips=chequearPAMIconTIPO(documento)
    sumar=botPUCOSUMAR(documento)
    return listaZips,sumar
def botPUCO2019(documento):
    global intentoListaDoc
    global nombreglobal
    nombreglobal = ""
    data = urllib.urlencode({'documento':documento,'tabla':configPOS['Puco']['Tabla']})
    try:
        url = configPOS['Puco']['UrlPrincipal']
        request = urllib2.Request(url,data)
        respuesta = urllib2.urlopen(request).read()
        respLimpia,nombreglobal=limpiandingSTR(respuesta)
        if respuesta == "null" or len(respLimpia) == 0:
            return "No se reportan datos"
        else:
            return respLimpia
    except:
        listaEr=["PAGINA NO DISPONIBLE"]
        return listaEr
#HACE UNA REQUEST Y DEVUELVE UN HTML
#FILTRAR VOMITO TE PARSEA TODO ESE HTML Y TE DEVUELVE EL NOMBRE, DOCUMENTO,... Y MAS IMPORTANTE TE DEVUELVE UN LINK
def chequearPAMI(documento):
	url = 'https://prestadores.pami.org.ar/result.php?c=6-2-2'
	data = urllib.urlencode({'tipoDocumento':'DNI','nro_doccumento':documento,'submit2':'Buscar'})
	request = urllib2.Request(url,data)
	respuesta = urllib2.urlopen(request).read()
	return filtrarVomitoV2(respuesta)
#MAS TE VALE NO ROMPER NADA HIJO DE PUTA------------------
global intentoListaDoc
global browser
global usuarioL
global contraL
global nombreglobal
global nombreglobal2
global nombreglobal3
global nombreglobalsumar
global configPOS
listaDoc=[]
listaNom=[]
listaOSES=[]
nombreglobal=""
nombreglobal2=""
nombreglobal3=""
configPOS = configparser.ConfigParser()
configPOS.sections()
configPOS.read('configPOS.ini')
usuarioL=configPOS['Superintendencia']['Usuario']
contraL=configPOS['Superintendencia']['Contra']
browser=Browser()
browser.set_handle_robots(False)
#PROBAR BOTS DE OBRAS SOCIALES -------------------------------------------------------------------
#print controlSPIN("40743779")
#print botPUCO2019("40743779")
#print botPUCOSUMAR("40743779")
#print chequearIOMA2018('40743779')
#PROBAR BOTS DE OBRAS SOCIALES -------------------------------------------------------------------
