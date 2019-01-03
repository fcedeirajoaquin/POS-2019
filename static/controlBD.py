import sqlite3
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
conn = sqlite3.connect("pos.sqlite3")
cursor = conn.cursor()
conn.text_factory = str
cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
print cursor.fetchall() 
# cursor.execute('''
#     CREATE TABLE historial(id INTEGER PRIMARY KEY, documento TEXT,
#                        nombre TEXT, puco TEXT , ioma TEXT , spin TEXT, fecha TEXT )
# ''')

# documento="40743779"
# nombre="Joaquin Fernandez Cedeira"
# puco="OBRA SOCIAL DE EMPRESARIOS, IOMA"
# ioma="TIENE IOMA"
# spin="OBRA SOCIAL DE EMPRESARIOS"
# users = [(name1,phone1, email1, password1),
#          (name2,phone2, email2, password2),
#          (name3,phone3, email3, password3)]
# cursor.execute('''INSERT INTO historial(documento,nombre,puco,ioma,spin)
#                   VALUES(:documento,:nombre, :puco, :ioma, :spin)''',
#                   {'documento':documento, 'nombre':nombre, 'puco':puco, 'ioma':ioma, 'spin':spin})
# cursor.execute('''DROP TABLE historial''')
# conn.commit()
cursor.execute('SELECT * FROM historial')
data = cursor.fetchall()
for elemento in data:
	print " "
 	print elemento

# PASAR A UN EXCEL----------------------------------------------------------------------------------------
# n=1
# i=1
# vacio=Workbook()
# hojavacia=vacio.active

# cursor.execute('SELECT * FROM historial')
# data = cursor.fetchall()
# for elemento in data:
# 	for item in elemento:
# 		hojavacia.cell(row=n,column=i).value=item
# 		i+=1
# 	n+=1
# 	i=1
# 	print " "
#  	print elemento
# vacio.save("BaseDeDatosPOS.xlsx")
# PASAR A UN EXCEL----------------------------------------------------------------------------------------
